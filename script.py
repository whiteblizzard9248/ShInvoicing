import openpyxl

wb = openpyxl.load_workbook('data/UKAdvertisers.xlsx')
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        if cell.value is not None:
            headers.append(cell.value)
    print("Headers:", headers)
    # Print first 5 rows
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= 5:
            break
        print(row)